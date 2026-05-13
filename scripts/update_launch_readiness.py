"""Add or update the Launch Readiness sheet in ARL_Test_Tracker.xlsx with
the 4 security-hardening fixes from docs/security_audit_2026-05-13.md.

Run: py scripts/update_launch_readiness.py
"""
import sys
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Force UTF-8 stdout on Windows so we never trip over a rupee/em-dash in output.
sys.stdout.reconfigure(encoding="utf-8")

WB_PATH = Path(__file__).resolve().parents[1] / "ARL_Test_Tracker.xlsx"
SHEET_NAME = "Launch Readiness"

HEADER = [
    "ID",
    "Area",
    "Item",
    "Severity",
    "Status",
    "Commit",
    "Last Tested",
    "Audit Reference",
    "Notes",
]

ROWS = [
    [
        "LR-SEC-002",
        "Security / DB",
        "S-002 — projects.latitude/longitude leak via marketplace select(); fixed via projects_public SECURITY INVOKER view (migration 034) + 4 repo repoints + model docstring",
        "High",
        "Done",
        "980776c",
        date(2026, 5, 13),
        "docs/security_audit_2026-05-13.md S-002",
        "Operator must run: supabase db push (applies 034). Verify with information_schema query in ops doc Part 7 §7.6.",
    ],
    [
        "LR-SEC-003",
        "Security / DB",
        "S-003 — public.sync_status flipped to SECURITY INVOKER (migration 035); same posture migration 015 took for portfolio_summary",
        "High",
        "Done",
        "0dd6225",
        date(2026, 5, 13),
        "docs/security_audit_2026-05-13.md S-003",
        "Operator must run: supabase db push. Confirm Studio → Advisors clears the security_definer_view warning.",
    ],
    [
        "LR-SEC-004",
        "Security / DB",
        "S-004 — REVOKE anon CRUD on post-019 tables (sync_alerts, user_settings, login_events, consultation_requests, exit_requests) via migration 036; defense-in-depth on top of RLS",
        "High",
        "Done",
        "91e03b0",
        date(2026, 5, 13),
        "docs/security_audit_2026-05-13.md S-004",
        "Operator must run: supabase db push. Verify with role_table_grants query in ops doc Part 7 §7.6 (only app_config/SELECT may remain).",
    ],
    [
        "LR-SEC-005",
        "Security / Edge Functions",
        "S-005 — replace wildcard Access-Control-Allow-Origin:* with APP_ALLOWED_ORIGINS allow-list via shared _shared/cors.ts; 5 edge functions repointed (bank-change-request, create-ticket, onboard-investor, reply-ticket, zoho-crm-webhook)",
        "High",
        "Done",
        "e279e94",
        date(2026, 5, 13),
        "docs/security_audit_2026-05-13.md S-005",
        "Operator must run: supabase secrets set APP_ALLOWED_ORIGINS=… then supabase functions deploy <name> for each of the 5 functions. CORS smoke test in ops doc Part 7 §7.7.",
    ],
]


def ensure_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        print(f"sheet '{SHEET_NAME}' exists; appending rows. existing rows={ws.max_row}")
        return ws, False
    ws = wb.create_sheet(SHEET_NAME)
    print(f"created sheet '{SHEET_NAME}'")
    return ws, True


def write_header(ws):
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="3C5152")  # ARL primary
    for col_idx, label in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    widths = [12, 22, 80, 10, 10, 12, 14, 36, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def upsert_rows(ws, fresh: bool):
    # Build id -> row map for existing rows so re-runs update in place.
    existing = {}
    if not fresh:
        for r in range(2, ws.max_row + 1):
            id_val = ws.cell(row=r, column=1).value
            if id_val:
                existing[str(id_val).strip()] = r

    next_row = ws.max_row + 1 if not fresh else 2
    if fresh:
        write_header(ws)

    for row in ROWS:
        rid = row[0]
        target = existing.get(rid, next_row)
        if target == next_row:
            next_row += 1
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=target, column=col_idx, value=val)
            if col_idx == 3 or col_idx == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Severity colour
        sev = row[3]
        sev_cell = ws.cell(row=target, column=4)
        sev_cell.fill = PatternFill(
            "solid",
            fgColor={"Critical": "C05640", "High": "D4AF37", "Medium": "E1DFC6"}.get(sev, "FFFFFF"),
        )
        # Status pill
        ws.cell(row=target, column=5).fill = PatternFill("solid", fgColor="2E7D6E")
        ws.cell(row=target, column=5).font = Font(color="FFFFFF", bold=True)
        ws.cell(row=target, column=5).alignment = Alignment(horizontal="center")
        # Date format
        ws.cell(row=target, column=7).number_format = "yyyy-mm-dd"
        print(f"  wrote {rid} at row {target}")


def main():
    if not WB_PATH.exists():
        print(f"ERROR: {WB_PATH} not found", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(WB_PATH)
    ws, fresh = ensure_sheet(wb)
    upsert_rows(ws, fresh)
    wb.save(WB_PATH)
    print(f"saved {WB_PATH}")


if __name__ == "__main__":
    main()
