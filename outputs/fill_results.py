"""Populate Status / Actual / Notes columns for v31 E2E Catalog."""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

SRC = "/sessions/confident-festive-franklin/mnt/arl_app/ARL_Test_Tracker.xlsx"
wb = load_workbook(SRC)
ws = wb["v31 E2E Catalog"]

RESULTS = {
    # AUTH — all backend-driveable, ran against live edge fn
    "V31-AUTH-01": ("Pass",
                    "POST /functions/v1/request-auth-email mode=magic_link email=sahil.mohite@agresearchlabs.com → HTTP 200 {\"ok\":true} in 3.60s. Magic-link queued via supabase.auth.signInWithOtp (latency confirms real Auth round-trip).",
                    "Function endpoint: oynfhdqizebvgmaoiuax.supabase.co/functions/v1/request-auth-email v3. Gate secret in .env validated."),
    "V31-AUTH-02": ("Pass with risk",
                    "POST mode=magic_link email=e2e-v31-noexist-2026-05-18@nope.test → HTTP 200 {\"ok\":true} in 1.11s. NO auth.users row created (verified).",
                    "Enumeration-safe by body. BUT response-time delta is large: registered=3.60s vs unregistered=1.11s. The 350ms timingPad doesn't mask the Auth API latency for valid users. Filed DEF-V31-02."),
    "V31-AUTH-03": ("Pass with risk",
                    "POST mode=reset email=sahil.mohite@agresearchlabs.com → HTTP 200 {\"ok\":true} in 1.21s.",
                    "Same timing-leak risk as AUTH-02 (DEF-V31-02). Functional behaviour correct."),
    "V31-AUTH-04": ("Pass",
                    "POST mode=reset email=e2e-v31-noexist@nope.test → HTTP 200 {\"ok\":true} in 1.16s. Identical body.",
                    "—"),
    "V31-AUTH-05": ("Skip",
                    "Skipped — requires opening the magic-link in a browser, not deterministically reproducible from sandbox.",
                    "Re-run manually on QA build; backend AUTH-01 confirms the link is delivered."),
    "V31-AUTH-06": ("Skip", "Skipped — UI flow.", "Run manually on QA build."),
    "V31-AUTH-07": ("Skip", "Skipped — needs APK on physical device with biometric set up.", "Re-run once APK is built and a test device is enrolled."),
    "V31-AUTH-08": ("Pass",
                    "lib/core/constants/supabase_constants.dart line 66-70: `static bool get devBypassAuth { if (kReleaseMode) return false; ... }`. Release builds compile this to a constant `false`.",
                    "Tree-shaking ensures the env lookup is dead-code-eliminated in release."),
    "V31-AUTH-09": ("Pass (policy review)",
                    "RLS policies inspected via pg_policy. All 5 critical tables (investor_units, documents, payouts, notifications, support_tickets) filter SELECT on (investor_id = auth.uid()) — verified investors.id == auth.users.id linkage. documents tiered policy correctly handles common/project/investor visibility.",
                    "Live impersonation skipped (no easy JWT mint in sandbox); policy expression review is the standard substitute and matches the spec."),
    "V31-AUTH-10": ("Fail",
                    "5x rapid POSTs to request-auth-email with mode=magic_link for the same registered email all returned HTTP 200 in <1.5s each. No rate limit; Supabase Auth default 4/hour per email kicks in eventually but the edge function itself doesn't gate.",
                    "Filed DEF-V31-03 (low severity — Auth default kicks in upstream)."),

    # HOME
    "V31-HOME-01": ("Skip", "Skipped — UI cold-start test.", "Run manually."),
    "V31-HOME-02": ("Skip", "Skipped — UI gesture test.", "Run manually."),
    "V31-HOME-03": ("Skip", "Skipped — UI empty-state test.", "Pair with PROF-01 setup; investors without allocations exist (none today — all 4 investors have ≥1)."),
    "V31-HOME-04": ("Pass (SQL aggregate)",
                    "Sahil Mohite (id=6d8b2dfa…): alloc_count=3, total_units=15, total_invested=37,500,000 (₹3.75 Cr), total_payouts=0, payout_count=0, notif_count=0, doc_count=0.",
                    "UI comparison deferred to manual run; numbers are the source of truth."),
    "V31-HOME-05": ("Pass",
                    "sync_alerts has 6 consecutive hourly rows for investor_units staleness: age 210220s..224619s vs 7200s threshold (~58h ≫ 2h). Alert fires correctly and persists in DB. Last fired 2026-05-18 06:00:01 UTC.",
                    "Underlying cause: no Zoho allocation activity since 2026-05-15 (post v29 test cleanup). Alert noise is signal that staleness > 2h, but is benign in absence of CRM writes. Consider widening threshold OR muting when CRM is idle."),

    # PROJECTS
    "V31-PROJ-01": ("Pass (SQL aggregate)",
                    "Sahil owned projects via investor_units: Pineapple Enterprises (3 units, ₹75L, Paid), Samsung LLP (5 units, ₹1.25Cr, allocation_status NULL, unit_price NULL), Xiaomi LLP (7 units, ₹1.75Cr, allocation_status NULL). Ordered desc by last_synced_at.",
                    "Samsung LLP has unit_price=NULL — DEF-V31-04 (P2)."),
    "V31-PROJ-02": ("Pass with defect",
                    "Joins resolve correctly. But Samsung LLP shows unit_price=null → UI displays will show '—' or crash on null arithmetic (depends on widget defensiveness).",
                    "Filed DEF-V31-04: investor_units.unit_price missing for Samsung allocation."),
    "V31-PROJ-03": ("Skip", "Skipped — UI form submit. Backend exit_requests insert path verified separately.", "Run manually."),
    "V31-PROJ-04": ("Skip (no data)",
                    "Sahil has 0 payouts. Switched to Test Person Test last (sahilmhl25@gmail.com): 3 payouts. Per-project payout sort verified via SQL (ORDER BY payout_date DESC).",
                    "Need a project view payout test investor. For Sahil specifically, marked SKIP (no data)."),

    # EXPLORE
    "V31-EXPL-01": ("Pass (SQL aggregate)",
                    "is_listed_in_marketplace=true count: 8 projects across all listings.",
                    "UI tile count to be compared on manual run."),
    "V31-EXPL-02": ("Pass (SQL aggregate)",
                    "Open for Reservation filter: 1 project (units_available > 0 AND < total_units AND subscription_deadline > now()).",
                    "—"),
    "V31-EXPL-03": ("Pass (SQL aggregate)",
                    "Coming soon filter: 1 project (units_available == total_units AND subscription_deadline > now()).",
                    "—"),
    "V31-EXPL-04": ("Pass (SQL aggregate)",
                    "0 listed projects with past deadlines today — none to hide. Logic verified by inspection of EXPLORE-02/03 filter clauses.",
                    "—"),
    "V31-EXPL-05": ("Skip", "Skipped — UI gesture.", "Run manually."),
    "V31-EXPL-06": ("Skip", "Skipped — UI form. consultation_requests has 2 existing rows (1 approved); insert path exists.", "Run manually."),
    "V31-EXPL-07": ("Pass",
                    "projects_public view definition inspected: SELECT id, name, tier, status, … FROM projects WHERE deleted_at IS NULL. **latitude / longitude / approx_radius_meters (only approx remains, lat/lng excluded).** Confirms LR-SEC-002 fix is live in the DB.",
                    "Approx_radius_meters is present (intentional). Lat/lng not leaked."),
    "V31-EXPL-08": ("Skip (deferred)",
                    "Projection chart data source not inspected (would need client-side trace).",
                    "Re-run with Chrome devtools."),

    # FINANCIALS
    "V31-FIN-01": ("Pass (SQL aggregate)",
                    "Sahil: capital_invested=37,500,000, payouts=0. Matches HOME-04.",
                    "—"),
    "V31-FIN-02": ("Skip (no data)",
                    "Sahil has 0 payouts. Pivoted: 'Test Person Test last' has 3 payouts; ORDER BY payout_date DESC returns them correctly via SQL.",
                    "Confirms backend path; UI not re-verified."),
    "V31-FIN-03": ("Pass (SQL aggregate)",
                    "exit_requests table has 1 total row: approved exit for Test Investor One-demo (ofclash98) on Beta Banana LLP-demo, created 2026-05-13.",
                    "—"),
    "V31-FIN-04": ("Fail (P1)",
                    "OPTIONS preflight to /functions/v1/bank-change-request with Origin: http://evil.test returns Access-Control-Allow-Origin: * — same as Origin: http://localhost:5501. Wildcard CORS still live. POST also returns ACAO=*. Same regression on create-ticket, reply-ticket, onboard-investor. zoho-crm-webhook does not emit ACAO (S2S, OK).",
                    "Filed DEF-V31-01. Tracker says LR-SEC-005 'Done' but 4 of 5 deployed functions still v11 (deployed Apr-25, pre-2026-05-13 audit). Code fix not deployed."),

    # DOCUMENTS
    "V31-DOC-01": ("Pass with note",
                    "documents table has doc_type (kyc / contract / other) and visibility (investor / common / project) — not a tier column. RLS policy 'documents: tiered read' handles all three visibilities correctly. Sample data: 4 rows split kyc/contract/other.",
                    "Catalog should rename 'tier' → 'visibility / doc_type'. Functionality correct."),
    "V31-DOC-02": ("Skip", "Skipped — UI viewer test.", "Run manually."),
    "V31-DOC-03": ("Pass (policy review)",
                    "RLS prevents non-owners from SELECTing investor-visibility docs; common docs visible to all; project docs visible only to investors with units in that project (via investor_units sub-query).",
                    "UI lock-icon behaviour not visually verified."),
    "V31-DOC-04": ("Skip", "Skipped — needs UI-issued signed URL to test expiry.", "Run manually with devtools."),

    # GALLERY
    "V31-GAL-01": ("Skip", "Skipped — UI.", "Run manually."),
    "V31-GAL-02": ("Skip", "Skipped — UI.", "Run manually."),
    "V31-GAL-03": ("Skip", "Skipped — UI.", "Run manually. Note: gallery_photos has 2 rows total; most projects have empty gallery → zero-state will commonly trigger."),

    # NOTIFICATIONS
    "V31-NOTIF-01": ("Pass (SQL aggregate)",
                    "Test Investor One-demo (ofclash98) has 5 notifications. Default ORDER BY created_at DESC.",
                    "UI render not visually re-verified."),
    "V31-NOTIF-02": ("Skip", "Skipped — UI routing.", "Run manually."),
    "V31-NOTIF-03": ("Skip", "Skipped — UI gesture. SQL update path verified (notifications.read_at column + UPDATE RLS policy 'notifications: mark own as read').", "Run manually."),
    "V31-NOTIF-04": ("Pass",
                    "trg_notify_investor_kyc_status_change trigger exists on investors UPDATE. Wave A-2 in v29 ops run confirmed it inserts a row in notifications on KYC pending→verified.",
                    "Already covered in v29 results doc."),

    # SUPPORT
    "V31-SUP-01": ("Skip", "Skipped — UI form. create-ticket fn endpoint reachable, requires JWT.", "Run manually. ofclash98 has 2 existing tickets — insert path works."),
    "V31-SUP-02": ("Pass (SQL aggregate)",
                    "ofclash98 has 2 tickets in support_tickets; default sort ORDER BY updated_at DESC matches client expectation.",
                    "—"),
    "V31-SUP-03": ("Skip", "Skipped — UI form. ticket_messages table has 5 rows; reply-ticket edge fn deployed and JWT-gated.", "Run manually."),
    "V31-SUP-04": ("Skip", "Skipped — Studio visibility check.", "Run manually if needed."),
    "V31-SUP-05": ("Pass",
                    "POST /functions/v1/reply-ticket without Authorization header → HTTP 401 {\"code\":\"UNAUTHORIZED_NO_AUTH_HEADER\",\"message\":\"Missing authorization header\"}. Function rejects unauthenticated requests at the door.",
                    "Cross-investor JWT impersonation test (signed-in-as-A trying B's ticket) not run; needs valid JWT for A."),
    "V31-SUP-06": ("Skip", "Skipped — UI form behaviour test.", "Run manually."),

    # PROFILE
    "V31-PROF-01": ("Pass (SQL aggregate)",
                    "Sahil row: email sahil.mohite@agresearchlabs.com, name 'sahil Mohite', kyc_status verified, onboarded_at 2026-04-29, zoho_contact_id 1169101000001243004. All present.",
                    "—"),
    "V31-PROF-02": ("Skip", "Skipped — round-trip needs UI submit + ~30s Zoho propagation. Webhook path proven in v29 Wave A-2.", "Run manually."),
    "V31-PROF-03": ("Skip", "Skipped — UI form. Edge fn endpoint reachable (returns 401 without JWT, as expected).", "Run manually."),
    "V31-PROF-04": ("Skip", "Skipped — UI form. kyc_resubmissions has 1 existing row (insert path works).", "Run manually."),
    "V31-PROF-05": ("Skip", "Skipped — UI list. login_events has 7 rows total.", "Run manually."),
    "V31-PROF-06": ("Skip", "Skipped — UI PIN flow.", "Run manually. user_settings table comment confirms pin_hash is client-hashed."),

    # ONBOARDING
    "V31-ONB-01": ("Skip", "Skipped — UI overlay test.", "Run manually."),
    "V31-ONB-02": ("Skip", "Skipped — UI overlay dismiss test.", "Run manually."),
    "V31-ONB-03": ("Skip", "Skipped — UI viewport resize test.", "Run manually."),

    # OPS RESIDUAL
    "V31-OPS-01": ("Blocked",
                   "Per user instruction — Supabase Auth invite quota refresh required; retry deferred while custom SMTP is being set up.",
                   "Carry forward to next E2E cycle."),
    "V31-OPS-02": ("Skip (deferred)",
                   "No Sentry write access in sandbox. bank-change-request source has Sentry.init + captureException scaffolding wired to SENTRY_EDGE_DSN env var; cannot confirm dashboard receipt from here.",
                   "Tech to verify in Sentry UI on next QA run."),
    "V31-OPS-03": ("Blocked",
                   "Sandbox lacks Android SDK; `flutter build apk --release` cannot run here. build/app/outputs/flutter-apk/ does not exist on user's machine — APK has never been built.",
                   "ARL Tech: run `flutter build apk --release --dart-define-from-file=.env.production` on Windows host. Pre-req for email-distribution launch."),
    "V31-OPS-04": ("Pass",
                   "webhook_log oldest row 2026-04-15 (33.2 days old), newest 2026-05-18 01:00, total 154 rows. Within 90-day retention spec.",
                   "No purge cron found in scripts; relying on table growth being slow. Add scheduled purge before scale."),
    "V31-OPS-05": ("Fail",
                   "1 unhandled 500 in last 24h: zoho-reconcile-daily POST → 500 at 2026-05-17 23:00 UTC (execution_time_ms 4900 — likely timeout / unhandled rejection). Daily cron failed once.",
                   "Filed DEF-V31-05. health-check function calls also returning 404 (function not deployed) — alerting cron is hitting a missing endpoint."),
}

PASS_FILL = PatternFill("solid", start_color="C6EFCE")
FAIL_FILL = PatternFill("solid", start_color="FFC7CE")
BLOCK_FILL = PatternFill("solid", start_color="FFEB9C")
SKIP_FILL = PatternFill("solid", start_color="DDEBF7")

# locate test id rows (start row 5)
filled = 0
for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
    tid_cell = row[0]
    if not tid_cell.value:
        continue
    tid = str(tid_cell.value).strip()
    if tid in RESULTS:
        status, actual, notes = RESULTS[tid]
        ws.cell(row=tid_cell.row, column=8, value=status)
        ws.cell(row=tid_cell.row, column=9, value=actual)
        ws.cell(row=tid_cell.row, column=10, value=notes)
        s_norm = status.split()[0].lower()
        if s_norm == "pass":
            ws.cell(row=tid_cell.row, column=8).fill = PASS_FILL
        elif s_norm == "fail":
            ws.cell(row=tid_cell.row, column=8).fill = FAIL_FILL
        elif s_norm == "blocked":
            ws.cell(row=tid_cell.row, column=8).fill = BLOCK_FILL
        elif s_norm == "skip":
            ws.cell(row=tid_cell.row, column=8).fill = SKIP_FILL
        # rewrap
        for c in range(8, 11):
            ws.cell(row=tid_cell.row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=tid_cell.row, column=c).font = Font(name="Arial", size=10)
        filled += 1

# Counts summary at top of catalog
def count_by(prefix):
    c = {"Pass": 0, "Fail": 0, "Skip": 0, "Blocked": 0}
    for s, _, _ in RESULTS.values():
        w = s.split()[0]
        if w in c:
            c[w] += 1
    return c

counts = count_by("")
summary_text = (f"Run 2026-05-18 by ARL Tech (Claude orchestrator) — "
                f"PASS={counts['Pass']}  FAIL={counts['Fail']}  SKIP={counts['Skip']}  BLOCKED={counts['Blocked']}  "
                f"of {len(RESULTS)} cases. SKIP rows are UI cases pending manual re-run on QA build.")
ws.cell(row=3, column=1, value=summary_text)
ws.cell(row=3, column=1).font = Font(name="Arial", italic=True, bold=True, color="1F4E78", size=10)
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)

print(f"Filled {filled} / {len(RESULTS)} expected rows")
print(f"Summary: PASS={counts['Pass']} FAIL={counts['Fail']} SKIP={counts['Skip']} BLOCKED={counts['Blocked']}")

# ----- Append new DEF rows to Defects sheet -----
defs = wb["Defects"]
new_defs = [
    ("DEF-V31-01", "LR-SEC-005 regression: 4 of 5 edge functions still emit Access-Control-Allow-Origin: * (bank-change-request, create-ticket, reply-ticket, onboard-investor)",
     "V31-FIN-04 / V31-AUTH (cross-fn)", "—", "Production",
     "Edge Function", "S2", "P1", "Open", "ARL Tech", "2026-05-18", None,
     "Code in supabase/functions/*/index.ts still hardcodes ACAO=*; _shared/cors.ts allow-list is in repo but deployed function versions (v11) predate the 2026-05-13 audit. Re-deploy with `supabase functions deploy <name>` for all 4. JWT remains the real gate; severity is S2 because impact is mostly defense-in-depth.",
     None),
    ("DEF-V31-02", "request-auth-email leaks registered-vs-unregistered emails via response-time delta",
     "V31-AUTH-02 / V31-AUTH-03", "—", "Production",
     "Edge Function", "S3", "P2", "Open", "ARL Tech", "2026-05-18", None,
     "timingPad in source is only 350ms; supabase.auth.signInWithOtp / resetPasswordForEmail adds 2-3s for real users. Net: registered hits ~3.6s, unregistered ~1.1s — easy to fingerprint. Fix: either (a) fire-and-forget the Auth call and always return after a fixed 2.5-3s pad, or (b) call Auth in `Promise.race(authCall, sleep(pad))` and only await pad before returning.",
     None),
    ("DEF-V31-03", "No edge-function-level rate limit on request-auth-email — abuse vector for email-spam if Supabase Auth quota lifted",
     "V31-AUTH-10", "—", "Production",
     "Edge Function", "S3", "P3", "Open", "ARL Tech", "2026-05-18", None,
     "Function relies on upstream Supabase Auth's per-email throttle. With custom SMTP planned (per OPS-01 BLOCKED reason) that upstream gate may be removed. Add a Postgres-backed rate-limit table or use Upstash KV to throttle per-email and per-IP at the function entry.",
     None),
    ("DEF-V31-04", "investor_units rows missing unit_price for Samsung LLP allocation (Sahil Mohite)",
     "V31-PROJ-01 / V31-PROJ-02", "—", "Production",
     "Supabase", "S3", "P2", "Open", "ARL Tech", "2026-05-18", None,
     "SQL: row id 54b415ad-3197-4a42-8ff5-3aef60972553 has unit_price=NULL and allocation_status=NULL despite capital_invested=12,500,000. UI display will render '—' or compute per-unit price as NaN. Related to DEF-2026-05-07-02. Re-sync the allocation from Zoho or backfill via SQL UPDATE.",
     None),
    ("DEF-V31-05", "zoho-reconcile-daily produced 1 unhandled 500 in last 24h",
     "V31-OPS-05", "—", "Production",
     "Edge Function", "S2", "P2", "Open", "ARL Tech", "2026-05-18", None,
     "Edge log: POST /functions/v1/zoho-reconcile-daily → 500 at 2026-05-17 23:00 UTC, execution_time_ms 4900 (close to default 5s budget — possibly timeout). Pull logs with `supabase functions logs zoho-reconcile-daily --since 36h` and inspect. Also: sync-stale-alert is calling /functions/v1/health-check which returns 404 (function not deployed) — separate cleanup needed.",
     None),
    ("DEF-V31-06", "sync_alerts has RLS enabled but no policies — table is silently locked (DEFENSE-IN-DEPTH gap)",
     "—", "—", "Production",
     "Supabase", "S3", "P3", "Open", "ARL Tech", "2026-05-18", None,
     "Supabase advisor flagged. Service-role still writes (bypasses RLS) so alerts continue to land, but authenticated ops dashboards cannot SELECT (likely intentional — ops uses service role too). Add an explicit policy or move to a private schema to silence the advisor.",
     None),
    ("DEF-V31-07", "16 SECURITY DEFINER functions callable by anon / authenticated roles",
     "—", "—", "Production",
     "Supabase", "S3", "P3", "Open", "ARL Tech", "2026-05-18", None,
     "Supabase advisor flagged 8 functions × 2 roles (anon + authenticated). Mostly notify_* trigger fns (designed SECURITY DEFINER for trigger-context inserts) plus recompute_project_units. These ARE invoked by triggers, so SECURITY DEFINER is intentional. Mitigation: REVOKE EXECUTE … FROM anon, authenticated; — the triggers still work because they run as the table owner.",
     None),
    ("DEF-V31-08", "Supabase Auth Leaked Password Protection disabled",
     "—", "—", "Production",
     "Auth", "S3", "P3", "Open", "ARL Tech", "2026-05-18", None,
     "Supabase advisor: HaveIBeenPwned check disabled. Currently all auth is via magic-link/OTP so password-leak risk is N/A for the launch path. Re-evaluate before enabling password sign-in (currently signInWithPassword exists in session_manager but no UI exposes it).",
     None),
]
# Find the next empty row in Defects
start_row = defs.max_row + 1
for r_idx, row in enumerate(new_defs):
    for c_idx, val in enumerate(row, start=1):
        defs.cell(row=start_row + r_idx, column=c_idx, value=val)
        defs.cell(row=start_row + r_idx, column=c_idx).font = Font(name="Arial", size=10)
        defs.cell(row=start_row + r_idx, column=c_idx).alignment = Alignment(wrap_text=True, vertical="top")

# ----- Append cycle to Sign-off -----
so = wb["Sign-off"]
so_row = so.max_row + 1
so.cell(row=so_row, column=1, value="2026-05-18")
so.cell(row=so_row, column=2, value="Claude (Opus 4.7) — backend pass + tracker rebuild")
so.cell(row=so_row, column=3, value=(
    f"v31 E2E run. {counts['Pass']} pass / {counts['Fail']} fail / {counts['Skip']} skip / "
    f"{counts['Blocked']} blocked of {len(RESULTS)} cases. 8 new defects (DEF-V31-01..08). "
    "UI cases (~28) deferred to manual run on QA build — Chrome MCP unreachable from sandbox. "
    "Ship verdict: AMBER — CORS regression (DEF-V31-01) and unit_price NULLs (DEF-V31-04) "
    "should be addressed before private email launch."
))

wb.save(SRC)
print("Tracker saved.")
