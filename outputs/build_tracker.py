"""Add Feature Backlog v1.1 + v31 E2E Catalog sheets to the tracker.
Idempotent — drops the sheets first if they exist, then re-creates."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/sessions/confident-festive-franklin/mnt/arl_app/ARL_Test_Tracker.xlsx"
wb = load_workbook(SRC)

# ---- common styles ----
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER = Border(*[Side(border_style="thin", color="C0C0C0")]*4)

# ===== Feature Backlog v1.1 =====
SHEET_FB = "Feature Backlog v1.1"
if SHEET_FB in wb.sheetnames:
    del wb[SHEET_FB]
fb = wb.create_sheet(SHEET_FB)

fb["A1"] = "Growize — Feature Backlog v1.1"
fb["A1"].font = TITLE_FONT
fb.merge_cells("A1:I1")
fb["A2"] = "Filed 2026-05-18 during v31 E2E pass. To be sequenced post-launch."
fb["A2"].font = SUBTITLE_FONT
fb.merge_cells("A2:I2")

fb_cols = ["ID", "Title", "Owner", "Size", "Priority", "Area", "Summary", "Dependencies / schema", "Notes"]
for i, h in enumerate(fb_cols, 1):
    c = fb.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER

fb_rows = [
    [
        "FB-V1.1-01",
        "WhatsApp RM button on Support screen",
        "TBD",
        "1-2 h",
        "P1",
        "Support",
        "Replace Support screen primary CTA with a deep-link to wa.me/<RM_PHONE> prefilled with investor name + ARL ID. Ticket form moves to secondary link.",
        "Add env var arlRmWhatsappNumber. No DB change. v1.2: make data-driven via investors.assigned_rm_phone.",
        "HNI investors prefer WhatsApp over forms — repeated qualitative signal from soft-launch conversations.",
    ],
    [
        "FB-V1.1-02",
        "Share-project button (Explore) + post-payout referral nudge",
        "TBD",
        "2-3 d",
        "P1",
        "Explore / Growth",
        "Per-project share icon in Explore + project detail. Tap → mint-share-token edge fn → native share sheet with wa.me / deep-link. Post-payout bottom-sheet nudges referral.",
        "New table share_tokens(token, project_id, referrer_investor_id, expires_at, clicked_count). Add investors.referred_by FK. New edge fn mint-share-token. Deep-link route /p/<token>.",
        "Free growth loop. Fills next project without ad spend. Attribution window 30 days from first click.",
    ],
    [
        "FB-V1.1-03",
        "First-payout celebration moment",
        "TBD",
        "1-2 d",
        "P0",
        "Home / Lifecycle",
        "Full-screen modal 'Your first payout has arrived — ₹X credited' with share/referral CTA. Triggers on home load when first payout exists and first_payout_seen=false.",
        "Add investors.first_payout_seen boolean default false. Lottie/animation asset. Reuses share infra from FB-V1.1-02.",
        "Highest-intent moment in investor lifecycle. Once-per-investor opportunity. Must land before first investor's first payout.",
    ],
]
for r_idx, row in enumerate(fb_rows, start=5):
    for c_idx, val in enumerate(row, start=1):
        c = fb.cell(row=r_idx, column=c_idx, value=val)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = BORDER

widths = [14, 36, 8, 8, 10, 18, 50, 50, 50]
for i, w in enumerate(widths, 1):
    fb.column_dimensions[get_column_letter(i)].width = w
for r in range(5, 5 + len(fb_rows)):
    fb.row_dimensions[r].height = 90

# ===== v31 E2E Catalog =====
SHEET_CAT = "v31 E2E Catalog"
if SHEET_CAT in wb.sheetnames:
    del wb[SHEET_CAT]
cat = wb.create_sheet(SHEET_CAT)

cat["A1"] = "Growize — v31 E2E Catalog (reusable test cases)"
cat["A1"].font = TITLE_FONT
cat.merge_cells("A1:J1")
cat["A2"] = ("Built 2026-05-18 against zoho-crm-webhook v31 + Flutter web @ localhost:5501 + Supabase oynfhdqizebvgmaoiuax. "
             "Re-run on every build by repeating Steps; record Pass/Fail/Skip/Blocked + Actual + Notes in the right-hand columns.")
cat["A2"].font = SUBTITLE_FONT
cat.merge_cells("A2:J2")

cat_cols = ["Test ID", "Area", "Description", "Steps", "Expected", "Tools", "Pre-reqs", "Status", "Actual", "Notes"]
for i, h in enumerate(cat_cols, 1):
    c = cat.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER

# ---- catalog cases ----
# columns: id, area, desc, steps, expected, tools, prereqs   (status/actual/notes filled at execution time)
CASES = [
    # A. AUTH
    ("V31-AUTH-01", "Auth", "Magic-link request for REGISTERED email",
     "1. POST /functions/v1/request-auth-email body={email: <registered>, action: 'magic_link'}\n2. Inspect response\n3. Check auth.users for invited_at",
     "200 generic 'If account exists, link sent'. Magic-link email queued (auth.users invited_at refreshed).",
     "curl + Supabase MCP SQL", "Registered investor exists (sahil.mohite@agresearchlabs.com)"),
    ("V31-AUTH-02", "Auth", "Magic-link request for UNREGISTERED email — enumeration-safe",
     "1. POST /functions/v1/request-auth-email body={email: <random>@nope.test, action: 'magic_link'}\n2. Inspect response code + body\n3. Confirm no auth.users row created",
     "Identical 200 response. NO row in auth.users. NO email sent.",
     "curl + Supabase MCP SQL", "—"),
    ("V31-AUTH-03", "Auth", "Forgot-password for REGISTERED email",
     "1. POST /functions/v1/request-auth-email body={email: <registered>, action: 'reset_password'}\n2. Inspect response",
     "200 generic. Reset link queued.",
     "curl", "Registered investor exists"),
    ("V31-AUTH-04", "Auth", "Forgot-password for UNREGISTERED email",
     "1. POST /functions/v1/request-auth-email body={email: <random>@nope.test, action: 'reset_password'}\n2. Inspect response",
     "Identical 200 response. No email sent.",
     "curl", "—"),
    ("V31-AUTH-05", "Auth", "Sign-in via magic-link token",
     "1. Trigger magic link for sahil.mohite@agresearchlabs.com\n2. Follow link in browser\n3. Assert lands on home tab with session",
     "Session JWT present; home tab loads with cached portfolio.",
     "Chrome MCP / manual", "AUTH-01 passed"),
    ("V31-AUTH-06", "Auth", "Sign-out clears session",
     "1. Logged-in app\n2. Profile → Sign out\n3. Reload",
     "Lands on login screen; LocalStorage / Supabase session keys cleared.",
     "Chrome MCP / manual", "AUTH-05 passed"),
    ("V31-AUTH-07", "Auth", "Biometric re-auth restores session",
     "1. App with biometric enabled\n2. Lock device\n3. Resume → biometric prompt → success",
     "Session restored from cached refresh token without magic-link round-trip.",
     "Mobile APK / manual", "APK build + device biometric set"),
    ("V31-AUTH-08", "Auth", "Dev-bypass flag is FALSE in release build",
     "1. Inspect lib/env/env.dart or build config\n2. grep for kDevBypass / authBypass\n3. Confirm release-mode constant is false",
     "Release builds compile with bypass disabled; flag is not a runtime override.",
     "grep / Read", "—"),

    # B. HOME / DASHBOARD
    ("V31-HOME-01", "Home", "Cold-start with cached data",
     "1. Kill app\n2. Re-open while offline (or block network)\n3. Observe Home",
     "Last-known data renders immediately from local cache; banner indicates offline mode.",
     "Chrome MCP / manual", "Previously logged-in session"),
    ("V31-HOME-02", "Home", "Pull-to-refresh triggers sync badge",
     "1. Home loaded\n2. Pull down on scrollview\n3. Observe sync badge",
     "Badge shows 'Syncing…' then 'Live'; data refetched from Supabase.",
     "Chrome MCP / manual", "Active session"),
    ("V31-HOME-03", "Home", "Empty state when investor has zero allocations",
     "1. Sign in as an investor with 0 rows in investor_units\n2. Open Home",
     "Home renders empty-state copy 'No investments yet…'. No crash.",
     "Chrome MCP / Supabase SQL", "Investor without allocations available"),
    ("V31-HOME-04", "Home", "Totals on Home match SQL aggregate",
     "1. SQL: SUM(iu.issued_units * iu.unit_price) for sahil.mohite\n2. Open Home as sahil.mohite\n3. Compare total invested + units",
     "UI total invested equals SQL SUM; units equals SUM(issued_units).",
     "Supabase MCP SQL + Chrome MCP", "Sahil Mohite session"),

    # C. PROJECTS
    ("V31-PROJ-01", "Projects", "Owned projects list view",
     "1. Sign in\n2. Projects tab\n3. Count + ordering",
     "All projects investor has units in, ordered by allocation_date desc.",
     "Chrome MCP + SQL", "Investor with >=1 allocation"),
    ("V31-PROJ-02", "Projects", "Project detail card fields",
     "1. Tap any owned project\n2. Verify: name, location, units owned, total invested, expected return, payout history",
     "All six fields populated from projects + investor_units + payouts joins.",
     "Chrome MCP + SQL", "PROJ-01"),
    ("V31-PROJ-03", "Projects", "Exit-request form submits to exit_requests",
     "1. Project detail → Request Exit\n2. Fill required fields\n3. Submit\n4. SQL exit_requests for this investor_unit_id",
     "New exit_requests row inserted; status='pending'.",
     "Chrome MCP + SQL", "PROJ-02"),
    ("V31-PROJ-04", "Projects", "Payout history sorted desc by date",
     "1. Project detail → Payout history\n2. Note order",
     "Most recent payout first; matches SQL ORDER BY paid_at DESC.",
     "Chrome MCP + SQL", "Investor with payouts"),

    # D. EXPLORE
    ("V31-EXPL-01", "Explore", "Pill 'All' shows all marketplace projects",
     "1. Explore tab → 'All' pill\n2. Count tiles\n3. SQL count(*) projects where is_listed_in_marketplace",
     "Tile count equals SQL count of listed projects.",
     "Chrome MCP + SQL", "—"),
    ("V31-EXPL-02", "Explore", "Pill 'Open for Reservation' filter logic",
     "1. Tap 'Open for Reservation' pill\n2. SQL: projects WHERE units_available > 0 AND units_available < total_units AND subscription_deadline > now()",
     "UI tiles = SQL result set.",
     "Chrome MCP + SQL", "—"),
    ("V31-EXPL-03", "Explore", "Pill 'Coming soon' filter logic",
     "1. Tap 'Coming soon' pill\n2. SQL: projects WHERE units_available = total_units AND subscription_deadline > now()",
     "UI tiles = SQL result set.",
     "Chrome MCP + SQL", "—"),
    ("V31-EXPL-04", "Explore", "Past-deadline project hidden",
     "1. Pick a project with subscription_deadline < now()\n2. Verify across all pills",
     "Not visible in any pill.",
     "Chrome MCP + SQL", "—"),
    ("V31-EXPL-05", "Explore", "Project tile opens detail card with projection chart",
     "1. Tap a tile\n2. Detail card renders\n3. Projection chart populated",
     "Detail card opens with chart data; no fallback empty chart.",
     "Chrome MCP", "—"),
    ("V31-EXPL-06", "Explore", "Request consultation submits to consultation_requests",
     "1. Detail → Request consultation\n2. Submit\n3. SQL count consultation_requests where investor_id=<self>",
     "New row inserted; notify-consultation-request fires.",
     "Chrome MCP + SQL", "Authenticated session"),

    # E. FINANCIALS
    ("V31-FIN-01", "Financials", "Portfolio summary matches SQL",
     "1. Open Financials\n2. SQL: SUM(capital_invested), SUM(payouts.amount) for investor\n3. Compare",
     "UI numbers = SQL aggregates within rounding.",
     "Chrome MCP + SQL", "Sahil Mohite session"),
    ("V31-FIN-02", "Financials", "Payouts list sorted desc",
     "1. Financials → Payouts\n2. Compare to SQL ORDER BY paid_at DESC",
     "Matches order + amounts.",
     "Chrome MCP + SQL", "Investor with payouts"),
    ("V31-FIN-03", "Financials", "Exit history reflects exit_requests",
     "1. Financials → Exit history\n2. SQL exit_requests for this investor",
     "Count + statuses match.",
     "Chrome MCP + SQL", "—"),

    # F. DOCUMENTS
    ("V31-DOC-01", "Documents", "Documents grouped by tier (KYC/Legal/Project)",
     "1. Documents tab\n2. Verify section headers\n3. SQL: documents by tier for investor",
     "Three sections populated correctly.",
     "Chrome MCP + SQL", "—"),
    ("V31-DOC-02", "Documents", "PDF opens in in-app viewer (pdfx)",
     "1. Tap a PDF document\n2. Confirm in-app viewer (not browser tab)",
     "PDF renders inside the Flutter pdfx viewer. NO external browser handoff.",
     "Chrome MCP / manual", "Investor with documents"),
    ("V31-DOC-03", "Documents", "Restricted document shows lock icon",
     "1. Sign in as investor without access to a tier-restricted doc\n2. Find a doc they don't own\n3. Observe icon + tap behaviour",
     "Lock icon shown; tap does NOT preview the file.",
     "Chrome MCP + SQL", "—"),

    # G. GALLERY
    ("V31-GAL-01", "Gallery", "Photos load lazily from arl-gallery bucket",
     "1. Open a project → Gallery\n2. Scroll\n3. Network panel: requests issued only as items enter viewport",
     "Lazy load confirmed; signed URLs from Storage.",
     "Chrome MCP", "Project with gallery photos"),
    ("V31-GAL-02", "Gallery", "Tap photo → fullscreen with pinch-zoom",
     "1. Tap thumbnail\n2. Confirm fullscreen modal\n3. Pinch to zoom",
     "Fullscreen viewer with working pinch gesture.",
     "Chrome MCP / manual", "GAL-01"),
    ("V31-GAL-03", "Gallery", "Empty gallery zero-state",
     "1. Pick project with 0 gallery_photos\n2. Open Gallery",
     "Friendly empty-state copy. No crash.",
     "Chrome MCP + SQL", "Project with 0 photos exists"),

    # H. ACTIVITY / NOTIFICATIONS
    ("V31-NOTIF-01", "Notifications", "Activity feed loads recent notifications",
     "1. Activity tab\n2. SQL: notifications for investor\n3. Compare count + ordering",
     "Matches SQL; ordered by created_at desc.",
     "Chrome MCP + SQL", "Investor with notifications"),
    ("V31-NOTIF-02", "Notifications", "Tap routes to relevant screen",
     "1. Tap a payout-type notification\n2. Confirm lands on Financials\n3. Tap exit-approved notification → Project detail",
     "Routing matches type-specific deep-links.",
     "Chrome MCP", "Investor with mixed notification types"),
    ("V31-NOTIF-03", "Notifications", "Mark-as-read updates read_at",
     "1. SQL: SELECT read_at FROM notifications WHERE id=<x> AND investor_id=<self>\n2. View notification in app\n3. Re-query",
     "read_at transitions from NULL to a timestamp.",
     "Chrome MCP + SQL", "Unread notification available"),

    # I. SUPPORT
    ("V31-SUP-01", "Support", "Create-ticket validation + insert",
     "1. Support → New ticket\n2. Submit empty → validation\n3. Submit valid → create-ticket edge fn\n4. SQL: support_tickets",
     "Validation blocks empty submit. Valid submit creates row, returns 200.",
     "Chrome MCP + edge-fn curl + SQL", "Authenticated session"),
    ("V31-SUP-02", "Support", "Ticket list sorted by updated_at desc",
     "1. Support → Tickets\n2. Compare to SQL ORDER BY updated_at DESC",
     "Matches.",
     "Chrome MCP + SQL", "Investor with tickets"),
    ("V31-SUP-03", "Support", "Reply-ticket edge fn appends to thread",
     "1. Pick a ticket\n2. Send reply\n3. SQL: ticket_messages for this ticket",
     "New ticket_messages row inserted; updated_at on support_tickets refreshed.",
     "Chrome MCP + curl + SQL", "Open ticket exists"),
    ("V31-SUP-04", "Support", "App-created ticket visible to ops (Studio / Zoho)",
     "1. Create ticket [E2E-V31] from app\n2. Open Supabase Studio support_tickets table\n3. Confirm row visible to ops",
     "Ops sees the ticket with status open.",
     "Chrome MCP + SQL", "SUP-01"),

    # J. PROFILE
    ("V31-PROF-01", "Profile", "Profile shows name/email/phone/KYC from investors row",
     "1. Profile tab\n2. SQL: investors row for this user\n3. Compare every field",
     "All fields match.",
     "Chrome MCP + SQL", "Sahil Mohite session"),
    ("V31-PROF-02", "Profile", "Edit name mirrors back to Zoho via Push_Contact_Update",
     "1. Profile → Edit name → save\n2. SQL: investors.name updated\n3. Zoho Contact .Last_Name verified post webhook",
     "investors.name updated; Zoho Contact row reflects change within a minute.",
     "Chrome MCP + SQL + Zoho MCP", "—"),
    ("V31-PROF-03", "Profile", "Bank change request submits via bank-change-request edge fn",
     "1. Profile → Bank → Request change → submit\n2. SQL: bank_change_requests new row\n3. 7-day cooldown enforced on re-submit",
     "Row inserted with status pending; second submit within 7 days returns cooldown error.",
     "Chrome MCP + curl + SQL", "—"),
    ("V31-PROF-04", "Profile", "KYC re-submission flow",
     "1. Set investor.kyc_status='rejected' (admin)\n2. Profile → KYC → resubmit\n3. SQL: kyc_resubmissions inserted",
     "Row inserted; ops sees pending request.",
     "Chrome MCP + SQL", "—"),

    # K. ONBOARDING
    ("V31-ONB-01", "Onboarding", "Tutorial shows on first-time home load",
     "1. Set investors.onboarded_at = NULL for test investor\n2. Sign in → Home",
     "Tutorial overlay renders.",
     "Chrome MCP + SQL", "—"),
    ("V31-ONB-02", "Onboarding", "Dismiss sets onboarded_at and suppresses re-show",
     "1. From ONB-01, dismiss tutorial\n2. SQL: investors.onboarded_at NOT NULL\n3. Sign out / sign in",
     "Tutorial does not re-appear.",
     "Chrome MCP + SQL", "ONB-01 ran"),
    ("V31-ONB-03", "Onboarding", "Welcome screen renders at narrow + wide viewport",
     "1. Sign-out → /login\n2. Resize to 380px wide\n3. Resize to 1440px wide",
     "Both layouts render. Wide-viewport cosmetic bug noted but non-blocking.",
     "Chrome MCP resize", "—"),

    # L. OPS RESIDUAL
    ("V31-OPS-01", "Ops Residual", "DEF-V29-04 — Allocation payout slot-1 retry",
     "1. Trigger allocation create with UTR_1 populated\n2. Inspect payouts table for slot 1 row",
     "Slot 1 payout row present.",
     "Zoho MCP + SQL", "Supabase Auth invite quota refreshed"),
    ("V31-OPS-02", "Ops Residual", "Sentry — synthetic error reaches dashboard",
     "1. Trigger a deliberate throw in app (e.g., a debug button or a /__crash__ route)\n2. Confirm Sentry event id",
     "Event captured in Sentry dashboard within 60s.",
     "Sentry / manual", "Sentry write access"),
    ("V31-OPS-03", "Ops Residual", "Mobile APK build clean",
     "1. flutter build apk --release\n2. Check exit code + artifact path",
     "build/app/outputs/flutter-apk/app-release.apk exists; no analyzer or compile errors.",
     "bash", "Flutter + Android SDK installed"),

    # additions for breadth
    ("V31-AUTH-09", "Auth", "RLS denies cross-investor data access",
     "1. Authenticated as investor A\n2. SELECT investor_units WHERE investor_id != self\n3. SELECT documents WHERE investor_id != self",
     "Zero rows returned. RLS filters correctly.",
     "Supabase MCP SQL (impersonate)", "—"),
    ("V31-HOME-05", "Home", "Sync badge surfaces 'Stale' when last_synced_at > 30m old",
     "1. Update llps.last_synced_at = now() - interval '45m'\n2. Pull-to-refresh on Home",
     "Badge shows 'Stale' or warning state.",
     "Chrome MCP + SQL", "—"),
    ("V31-EXPL-07", "Explore", "Marketplace projects use projects_public view (no lat/lng leak)",
     "1. Network panel during Explore load\n2. Inspect response: latitude/longitude absent",
     "Lat/lng fields not in response (post LR-SEC-002 fix).",
     "Chrome MCP devtools / Supabase SQL", "Migration 034 applied"),
    ("V31-FIN-04", "Financials", "Edge function bank-change-request CORS allow-list",
     "1. curl with Origin: http://localhost:5501 → expect 200/204 preflight\n2. curl with Origin: http://evil.test → expect blocked CORS",
     "Allowed origin returns ACAO header; disallowed origin returns no/empty ACAO.",
     "curl", "APP_ALLOWED_ORIGINS env set"),
    ("V31-DOC-04", "Documents", "Signed URLs expire within configured window",
     "1. Tap document → grab signed URL from network\n2. Wait past expiry\n3. Re-fetch URL",
     "Expired URL returns 403/410.",
     "Chrome MCP devtools / curl", "—"),
    ("V31-SUP-05", "Support", "Reply edge fn rejects non-owner",
     "1. As investor A, attempt POST /functions/v1/reply-ticket with ticket_id owned by investor B\n2. Inspect response",
     "403 / not-owner error. No insert into ticket_messages.",
     "curl + SQL", "Two investors + ticket fixtures"),
    ("V31-NOTIF-04", "Notifications", "KYC verify trigger creates notification row",
     "1. SQL: UPDATE investors SET kyc_status='verified' for fixture\n2. SQL: SELECT notifications for fixture",
     "trg_notify_investor_kyc_status_change inserts a 'kyc' type row.",
     "Supabase MCP SQL", "Throwaway investor"),
    ("V31-PROF-05", "Profile", "Login history reads from login_events",
     "1. Profile → Security → Login history\n2. SQL: login_events for self",
     "UI list matches SQL rows (count + timestamps).",
     "Chrome MCP + SQL", "—"),
    ("V31-OPS-04", "Ops Residual", "webhook_log purge policy verified",
     "1. SQL: SELECT min(created_at) FROM webhook_log\n2. Confirm older than 90 days are pruned",
     "min(created_at) is within 90 days OR a documented cron exists.",
     "Supabase MCP SQL", "—"),
    ("V31-OPS-05", "Ops Residual", "Edge function logs free of 5xx in last 24h",
     "1. Supabase MCP get_logs service=edge-function last 24h\n2. Count 5xx",
     "Zero unhandled 500s across all 11 functions.",
     "Supabase MCP get_logs", "—"),
    ("V31-AUTH-10", "Auth", "Magic-link rate limit gating (anti-spam)",
     "1. POST request-auth-email 6x within 60s for same email\n2. Inspect responses",
     "After Nth call, function returns 429 or generic OK without sending email (no enumeration leak).",
     "curl", "—"),
    ("V31-EXPL-08", "Explore", "Projection chart data source verified",
     "1. Tap project detail\n2. Verify chart fed from projects.expected_payout_schedule or equivalent",
     "Chart populates from real DB column, not a stub.",
     "Chrome MCP + Read", "—"),
    ("V31-PROF-06", "Profile", "PIN hash never persisted plaintext",
     "1. Set a PIN in app\n2. SQL: user_settings.pin_hash for self\n3. Confirm hashed string, not plaintext",
     "pin_hash is SHA-256-style hex; plaintext PIN absent from DB.",
     "Chrome MCP + SQL", "—"),
    ("V31-SUP-06", "Support", "Resolved ticket blocks further replies",
     "1. Mark a ticket status='resolved' in Studio\n2. From app, attempt reply",
     "Reply form disabled OR reply-ticket returns 4xx with 'ticket closed' error.",
     "Chrome MCP + curl + SQL", "—"),
]

for r_idx, (tid, area, desc, steps, exp, tools, prereq) in enumerate(CASES, start=5):
    cat.cell(row=r_idx, column=1, value=tid)
    cat.cell(row=r_idx, column=2, value=area)
    cat.cell(row=r_idx, column=3, value=desc)
    cat.cell(row=r_idx, column=4, value=steps)
    cat.cell(row=r_idx, column=5, value=exp)
    cat.cell(row=r_idx, column=6, value=tools)
    cat.cell(row=r_idx, column=7, value=prereq)
    cat.cell(row=r_idx, column=8, value="Pending")
    cat.cell(row=r_idx, column=9, value="")
    cat.cell(row=r_idx, column=10, value="")
    for c_idx in range(1, 11):
        c = cat.cell(row=r_idx, column=c_idx)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = BORDER

widths_cat = [14, 14, 38, 50, 38, 22, 26, 12, 50, 50]
for i, w in enumerate(widths_cat, 1):
    cat.column_dimensions[get_column_letter(i)].width = w
for r in range(5, 5 + len(CASES)):
    cat.row_dimensions[r].height = 85
cat.freeze_panes = "A5"

# Save
OUT = "/sessions/confident-festive-franklin/mnt/arl_app/ARL_Test_Tracker.xlsx"
wb.save(OUT)
print(f"Saved. Sheets now: {wb.sheetnames}")
print(f"Catalog rows: {len(CASES)}  Backlog rows: {len(fb_rows)}")
